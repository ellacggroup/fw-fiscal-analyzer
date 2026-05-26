"""
Export endpoints: Excel and PDF report generation for a stored agenda analysis.

  GET /agendas/{upload_id}/export/excel   → .xlsx download
  GET /agendas/{upload_id}/export/pdf     → .pdf download
"""

import io
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from database import AgendaItem, AgendaUpload, get_db

router = APIRouter(prefix="/agendas", tags=["export"])


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def _export_excel(upload: AgendaUpload, items: list[AgendaItem]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    FILL = {
        "POSITIVE": "C6EFCE",
        "NEUTRAL":  "FFEB9C",
        "NEGATIVE": "FFC7CE",
        "UNKNOWN":  "D9D9D9",
    }
    FONT_COLOR = {
        "POSITIVE": "276221",
        "NEUTRAL":  "9C5700",
        "NEGATIVE": "9C0006",
        "UNKNOWN":  "595959",
    }

    wb = Workbook()

    # ── Summary sheet ────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.column_dimensions["A"].width = 22
    ws_sum.column_dimensions["B"].width = 50

    def _hrow(label, value):
        row = ws_sum.max_row + 1
        ws_sum.cell(row, 1, label).font = Font(bold=True)
        ws_sum.cell(row, 2, str(value) if value else "—")

    ws_sum.cell(1, 1, "Fort Worth Fiscal Impact Analysis").font = Font(bold=True, size=14)
    ws_sum.merge_cells("A1:B1")
    ws_sum.append([])
    _hrow("Agenda file", upload.filename)
    _hrow("Meeting date", upload.meeting_date or "Unknown")
    _hrow("Analysis date", datetime.utcnow().strftime("%Y-%m-%d"))
    _hrow("Total items", upload.item_count)

    ratings = [i.analysis.get("fiscal_impact_rating", "UNKNOWN") for i in items if i.analysis]
    for r in ("POSITIVE", "NEUTRAL", "NEGATIVE", "UNKNOWN"):
        _hrow(f"  {r.capitalize()}", ratings.count(r))

    ws_sum.append([])
    _hrow("Analysis engine", "Claude AI + Rule-based" if any(
        i.analysis and i.analysis.get("claude_available") for i in items
    ) else "Rule-based")

    # ── Detail sheet ────────────────────────────────────────────────
    ws = wb.create_sheet("Agenda Items")
    headers = [
        "Item #", "Section", "Category", "Title",
        "Fiscal Rating", "Risk Level", "Recurring?",
        "Year 1 Net ($)", "40-yr Net ($)", "R/C Ratio",
        "Claude Summary", "Key Concerns", "Methodology Notes",
    ]
    ws.append(headers)
    for col_i, _ in enumerate(headers, 1):
        cell = ws.cell(1, col_i)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="003087")  # FW blue
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for item in items:
        a = item.analysis or {}
        rating = a.get("fiscal_impact_rating", "UNKNOWN")
        risk = a.get("risk_level", "")
        recurring = a.get("is_recurring")
        recurring_str = "Yes" if recurring is True else ("No" if recurring is False else "?")
        concerns = "; ".join(a.get("key_concerns") or [])
        summary = a.get("claude_summary") or a.get("analysis_narrative") or ""

        row = [
            item.item_number or "",
            (item.analysis or {}).get("section", ""),
            item.category or "",
            item.title or "",
            rating,
            risk,
            recurring_str,
            a.get("year1_net_impact"),
            a.get("projection_40yr_net"),
            a.get("revenue_to_cost_ratio"),
            summary[:500] if summary else "",
            concerns,
            (a.get("caveats") or "")[:300],
        ]
        ws.append(row)
        row_idx = ws.max_row
        fill_color = FILL.get(rating, "D9D9D9")
        font_color = FONT_COLOR.get(rating, "000000")
        for col_i in range(1, len(headers) + 1):
            cell = ws.cell(row_idx, col_i)
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col_i == 5:  # Rating column
                cell.font = Font(bold=True, color=font_color)

    # Column widths
    widths = [8, 18, 22, 50, 14, 10, 10, 14, 14, 10, 60, 40, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# PDF export (reportlab)
# ---------------------------------------------------------------------------

def _export_pdf(upload: AgendaUpload, items: list[AgendaItem]) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    )

    FW_BLUE = colors.HexColor("#003087")
    FW_GOLD = colors.HexColor("#C8962E")

    RATING_COLORS = {
        "POSITIVE": colors.HexColor("#C6EFCE"),
        "NEUTRAL":  colors.HexColor("#FFEB9C"),
        "NEGATIVE": colors.HexColor("#FFC7CE"),
        "UNKNOWN":  colors.HexColor("#E8E8E8"),
    }

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=letter,
        leftMargin=0.75 * inch,
        rightMargin=0.75 * inch,
        topMargin=0.75 * inch,
        bottomMargin=0.75 * inch,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title", parent=styles["Title"],
        textColor=FW_BLUE, fontSize=18, spaceAfter=4,
    )
    subtitle_style = ParagraphStyle(
        "Subtitle", parent=styles["Normal"],
        textColor=colors.gray, fontSize=10, spaceAfter=12,
    )
    h2_style = ParagraphStyle(
        "H2", parent=styles["Heading2"],
        textColor=FW_BLUE, fontSize=12, spaceBefore=16, spaceAfter=4,
    )
    body_style = ParagraphStyle(
        "Body", parent=styles["Normal"],
        fontSize=9, leading=13, spaceAfter=4,
    )
    small_style = ParagraphStyle(
        "Small", parent=styles["Normal"],
        fontSize=8, leading=11, textColor=colors.gray,
    )

    story = []

    # Cover / header
    story.append(Paragraph("Fort Worth City Council", title_style))
    story.append(Paragraph("Fiscal Impact Analysis Report", title_style))

    meta_lines = [upload.filename or ""]
    if upload.meeting_date:
        meta_lines.append(f"Meeting date: {upload.meeting_date}")
    meta_lines.append(f"Report generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}")
    story.append(Paragraph(" · ".join(meta_lines), subtitle_style))

    # Summary table
    ratings = [i.analysis.get("fiscal_impact_rating", "UNKNOWN") for i in items if i.analysis]
    engine = "Claude AI + Rule-based" if any(
        i.analysis and i.analysis.get("claude_available") for i in items
    ) else "Rule-based"

    summary_data = [
        ["Total items", str(len(items)), "Engine", engine],
        ["Positive", str(ratings.count("POSITIVE")), "Neutral", str(ratings.count("NEUTRAL"))],
        ["Negative", str(ratings.count("NEGATIVE")), "Unknown", str(ratings.count("UNKNOWN"))],
    ]
    summary_table = Table(summary_data, colWidths=[1.2 * inch, 0.8 * inch, 1.2 * inch, 2.5 * inch])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), FW_BLUE),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE",   (0, 0), (-1, -1), 9),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ("FONTNAME",   (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME",   (2, 0), (2, -1), "Helvetica-Bold"),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 0.25 * inch))

    # Item list
    story.append(Paragraph("Agenda Item Analysis", h2_style))

    col_widths = [0.5 * inch, 1.1 * inch, 3.2 * inch, 0.95 * inch, 0.85 * inch]
    tbl_header = [
        Paragraph("<b>#</b>", body_style),
        Paragraph("<b>Category</b>", body_style),
        Paragraph("<b>Summary</b>", body_style),
        Paragraph("<b>Rating</b>", body_style),
        Paragraph("<b>Risk</b>", body_style),
    ]
    tbl_data = [tbl_header]
    tbl_styles = [
        ("BACKGROUND",  (0, 0), (-1, 0), FW_BLUE),
        ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, -1), 8),
        ("LEADING",     (0, 0), (-1, -1), 11),
        ("VALIGN",      (0, 0), (-1, -1), "TOP"),
        ("GRID",        (0, 0), (-1, -1), 0.3, colors.lightgrey),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING",(0, 0), (-1, -1), 4),
        ("TOPPADDING",  (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
    ]

    for row_i, item in enumerate(items, 1):
        a = item.analysis or {}
        rating = a.get("fiscal_impact_rating", "UNKNOWN")
        risk = a.get("risk_level", "—") or "—"
        summary = a.get("claude_summary") or a.get("analysis_narrative") or "—"
        # Truncate long summaries
        if len(summary) > 260:
            summary = summary[:257] + "…"
        title_text = (item.title or "")[:120]

        row = [
            Paragraph(str(item.item_number or row_i), small_style),
            Paragraph(item.category or "—", small_style),
            Paragraph(f"<b>{title_text}</b><br/>{summary}", small_style),
            Paragraph(rating.capitalize(), small_style),
            Paragraph(risk, small_style),
        ]
        tbl_data.append(row)

        fill = RATING_COLORS.get(rating, colors.white)
        bg_row = row_i  # header is row 0
        tbl_styles.append(("BACKGROUND", (0, bg_row), (-1, bg_row), fill))

    items_table = Table(tbl_data, colWidths=col_widths, repeatRows=1)
    items_table.setStyle(TableStyle(tbl_styles))
    story.append(items_table)

    story.append(Spacer(1, 0.2 * inch))
    story.append(Paragraph(
        "Methodology: Rule-based analysis uses Fort Worth fiscal parameters (property tax $0.7125/$100 AV, "
        "Fate TX 40-year R/C framework). Claude AI provides qualitative summaries. "
        "Estimates are informational only.",
        small_style,
    ))

    doc.build(story)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Route handlers
# ---------------------------------------------------------------------------

@router.get("/{upload_id}/export/excel")
def export_excel(upload_id: int, db: Session = Depends(get_db)):
    upload, items = _load(upload_id, db)
    content = _export_excel(upload, items)
    filename = f"fw-fiscal-{upload_id}.xlsx"
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{upload_id}/export/pdf")
def export_pdf(upload_id: int, db: Session = Depends(get_db)):
    upload, items = _load(upload_id, db)
    content = _export_pdf(upload, items)
    filename = f"fw-fiscal-{upload_id}.pdf"
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _load(upload_id: int, db: Session) -> tuple[AgendaUpload, list[AgendaItem]]:
    upload = db.query(AgendaUpload).filter(AgendaUpload.id == upload_id).first()
    if not upload:
        raise HTTPException(status_code=404, detail="Agenda not found")
    items = db.query(AgendaItem).filter(AgendaItem.upload_id == upload_id).all()
    return upload, items
